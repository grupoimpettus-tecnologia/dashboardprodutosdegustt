import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from degust_produtos import MARCAS_CONFIG, carregar_produtos_marca

st.set_page_config(
    page_title="Painel Produtos Degust",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

COLUNAS_EXIBICAO = [
    "marca",
    "codigoLoja",
    "unidade",
    "categoria",
    "codigoProduto",
    "produtoExibicao",
    "precoFormatado",
]

COLUNAS_DOWNLOAD = [
    "marca",
    "codigoLoja",
    "unidade",
    "categoria",
    "tipo",
    "codigoPrincipal",
    "codigoProduto",
    "produtoExibicao",
    "produto",
    "preco",
]

ROTULOS_COLUNAS = {
    "marca": "Marca",
    "codigoLoja": "Cód. Unidade",
    "unidade": "Unidade",
    "categoria": "Categoria",
    "tipo": "Tipo",
    "codigoPrincipal": "Cód. Principal",
    "codigoProduto": "Cód. Produto",
    "produtoExibicao": "Produto",
    "produto": "Produto (original)",
    "precoFormatado": "Preço",
    "preco": "Preço",
}

st.markdown(
    """
    <style>
    .main { padding: 0rem 1rem; }
    .stButton>button {
        width: 100%;
        background-color: #366092;
        color: white;
        font-weight: bold;
        border-radius: 5px;
        border: none;
        padding: 10px;
    }
    .stButton>button:hover { background-color: #2a4a72; }
    h1 { color: #2c3e50; text-align: center; }
    /* Hierarquia de produtos na coluna Produto */
    [data-testid="stDataFrame"] td {
        white-space: pre !important;
        font-family: Consolas, "Courier New", monospace;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def _agora_brasilia_str():
    try:
        from zoneinfo import ZoneInfo

        return datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        from datetime import timedelta, timezone

        return datetime.now(timezone(timedelta(hours=-3))).strftime("%d/%m/%Y %H:%M:%S")


def criar_excel(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column in ws.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 50)

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@st.cache_data(ttl=300)
def carregar_marca(marca: str) -> pd.DataFrame:
    linhas = carregar_produtos_marca(marca)
    if not linhas:
        return pd.DataFrame()
    return pd.DataFrame(linhas)


def _renomear_colunas(df: pd.DataFrame, colunas: list[str]) -> pd.DataFrame:
    existentes = [c for c in colunas if c in df.columns]
    out = df[existentes].copy()
    return out.rename(columns={c: ROTULOS_COLUNAS.get(c, c) for c in existentes})


def main():
    st.title("📦 Painel Produtos Degust")
    st.markdown(
        "<div style='text-align: center;'>"
        "<strong>Espetto · Mané · Buteco Seu Rufino · Bendito</strong>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    with st.sidebar:
        st.header("⚙️ Controles")
        if st.button("🔄 Atualizar Dados", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        st.markdown("---")
        st.header("🏪 Filtrar por Marca")
        filtro_todas = st.checkbox("Todas as Marcas", value=False)

        if filtro_todas:
            marcas_selecionadas = list(MARCAS_CONFIG.keys())
        else:
            marcas_selecionadas = [
                marca for marca in MARCAS_CONFIG if st.checkbox(marca, value=False, key=f"marca_{marca}")
            ]

        st.markdown("---")
        st.markdown(f"**📅 Última atualização:**  \n{_agora_brasilia_str()}")

    if not marcas_selecionadas:
        st.warning("⚠️ Selecione pelo menos uma marca para visualizar os produtos.")
        return

    frames = []
    for marca in marcas_selecionadas:
        with st.spinner(f"Carregando produtos de {marca}…"):
            df_marca = carregar_marca(marca)
            if not df_marca.empty:
                frames.append(df_marca)

    if not frames:
        st.error("❌ Nenhum produto encontrado para as marcas selecionadas.")
        return

    df = pd.concat(frames, ignore_index=True)

    st.markdown("### 🔍 Filtros")
    c1, c2, c3 = st.columns(3)

    with c1:
        opcoes_unidades = sorted(df["unidade"].dropna().unique().tolist())
        unidades_sel = st.multiselect(
            "Unidades",
            options=opcoes_unidades,
            default=[],
            placeholder="Todas as unidades",
        )

    with c2:
        opcoes_categorias = sorted(df["categoria"].dropna().unique().tolist())
        categorias_sel = st.multiselect(
            "Categorias",
            options=opcoes_categorias,
            default=[],
            placeholder="Todas as categorias",
        )

    with c3:
        busca = st.text_input("Buscar produto ou código", placeholder="Ex: Chopp ou 1234")

    df_filtrado = df.copy()
    if unidades_sel:
        df_filtrado = df_filtrado[df_filtrado["unidade"].isin(unidades_sel)]
    if categorias_sel:
        df_filtrado = df_filtrado[df_filtrado["categoria"].isin(categorias_sel)]
    if busca.strip():
        termo = busca.strip().lower()
        df_filtrado = df_filtrado[
            df_filtrado["produto"].astype(str).str.lower().str.contains(termo, na=False)
            | df_filtrado.get("produtoExibicao", df_filtrado["produto"]).astype(str).str.lower().str.contains(termo, na=False)
            | df_filtrado["codigoProduto"].astype(str).str.contains(termo, na=False)
        ]

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("📦 Produtos", len(df_filtrado))
    with m2:
        st.metric("🏪 Unidades", df_filtrado["unidade"].nunique())
    with m3:
        st.metric("🏷️ Categorias", df_filtrado["categoria"].nunique())
    with m4:
        st.metric("🎯 Códigos únicos", df_filtrado["codigoProduto"].nunique())

    st.markdown("---")

    df_exibir = _renomear_colunas(df_filtrado, COLUNAS_EXIBICAO)
    st.dataframe(df_exibir, use_container_width=True, height=650)

    df_download = _renomear_colunas(df_filtrado, COLUNAS_DOWNLOAD)
    st.download_button(
        label="⬇️ Download Excel",
        data=criar_excel(df_download),
        file_name=f"painel_produtos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
